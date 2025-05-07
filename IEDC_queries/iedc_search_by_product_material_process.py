# -*- coding: utf-8 -*-
"""
Created on Tue Jul 25 20:17:49 2023

@author: spauliuk

SQL commands and Python code for iedc search for given datatypes by product/commodity, material and process
"""

import pymysql
import datetime
import IEDC_PW
import openpyxl
import numpy as np

conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

no_groups    = 17 # number of data types for drop-down menu
data_label  = []
data_type   = []
data_aspect = []
# fetch input data:
mywb = openpyxl.load_workbook('IEDC_product_material_search_vMay2025.xlsx')
for mr in range(0,no_groups):
    data_label.append(mywb['Data_type_menu_entries'].cell(mr+3,2).value)
    data_type.append(mywb['Data_type_menu_entries'].cell(mr+3,9).value)
    data_aspect.append(mywb['Data_type_menu_entries'].cell(mr+3,11).value)
    
# prepare export table with labels
book = openpyxl.Workbook() # Compile query results
ws1 = book.active
ws1.title = 'lookup_values'  
ws1.cell(row=1, column=1).value = 'Dropdown labels'
ws1.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    
for mr in range(0,no_groups):
    dsilist = []
    dsnlist = []
    dsalist = []
    mcur    = 0
    cur.execute("SELECT * FROM datasets WHERE data_type = %s",(str(data_type[mr])))
    for row in cur:
        print(row)
        dsilist.append(row[0]) # dataset_name
        dsnlist.append(row[1]) # dataset_name
        dsalist.append([])
        for ma in range(0,12):
            dsalist[mcur].append(row[22+2*ma])
        mcur += 1
    
    cimatch = []    
    for ma in range(0,12): # all aspects from 1 to 12
        for mcur in range(0,len(dsnlist)):    # for all datasets in the match
            if dsalist[mcur][ma] == data_aspect[mr]:
                print(mcur,ma)
                #SQL  = "SELECT aspect" + str(ma+1) + " FROM data WHERE dataset_id = " + str(dsilist[mcur])
                #cur.execute(SQL)
                SQL1 = "SELECT attribute1_oto FROM classification_items JOIN data ON classification_items.id = data.aspect" + str(ma+1) + " WHERE dataset_id = " + str(dsilist[mcur])
                cur.execute(SQL1)
                for row in cur:
                    cimatch.append(row[0])
            
    no_hits     = len(cimatch)
    searchitems = sorted(list(set(cimatch)), key=str.casefold)
       
    ws1.cell(row=2+mr, column=1).value = data_label[mr] +" (" + str(no_hits) + " entries)"
    ws1.cell(row=1, column=2+mr).value = 'Labels_' + str(mr)
    ws1.cell(row=1, column=2+mr).font = openpyxl.styles.Font(bold=True)
    for mrs in range(0,len(searchitems)):
        ws1.cell(row=2+mrs, column=2+mr).value = searchitems[mrs]
      
book.save('lookupvalues.xlsx')

#        