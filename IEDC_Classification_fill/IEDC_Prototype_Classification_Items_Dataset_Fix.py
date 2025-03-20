# -*- coding: utf-8 -*-
"""
Created on Thu Mar 20 13:10:15 2025

@author: Stefan Pauliuk

The classifiation items table is case insensitive, meaning that
if 'copper' is already included, 'Copper' cannot be added.
This is actually good for us, since we don't want to create redundant labels
that just differ in the way the items are capitalized.
However, this means that a label "Copper" will always be rejected.
The solution: A script that changes a list of labels to those
that are either inserted or already present.'
"""

import pymysql
import openpyxl

import IEDC_PW
import IEDC_Paths

# open iedc connection:
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')
cur = conn.cursor() 

# open workbook with labels:
mywb    = openpyxl.load_workbook('IEDC_Validation_Labels_Fix_v1.xlsx')
this_ID = mywb['Labels'].cell(1,2).value
these_L = []
rind    = 2
while True:
    if mywb['Labels'].cell(rind,2).value is None:
        break
    else:
        these_L.append(mywb['Labels'].cell(rind,2).value)
        rind += 1

for li in range(0,len(these_L)):
    try: 
        cur.execute("INSERT INTO classification_items (classification_id,attribute1_oto) VALUES (%s,%s)",(this_ID,these_L[li]))
        mywb['Labels'].cell(li+2,3).value = these_L[li]
    except: # value is rejected, already exists, fetch it:
        cur.execute("SELECT attribute1_oto FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(this_ID,these_L[li]))    
        for row in cur:
            mywb['Labels'].cell(li+2,3).value = row[0]

#save results            
mywb.save('IEDC_Validation_Labels_Fix_v1_filled.xlsx')            

# Close connection
cur.close()
conn.close()