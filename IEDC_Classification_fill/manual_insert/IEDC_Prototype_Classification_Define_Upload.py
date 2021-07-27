# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

@author: spauliuk
"""

import pymysql
import datetime
import numpy as np
from os import walk, path
import openpyxl
from datetime import date, timedelta
import csv

import IEDC_PW
import IEDC_Paths

def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1


ClassList   = ['IEDC_73_MFA13'] # List of filenames for classifications to be added.
ClassIDList = [73]


# Define mySQL commands for classification
SQLD = "INSERT INTO classification_definition (\
id,\
classification_name,\
dimension,\
description,\
mutually_exclusive,\
collectively_exhaustive,\
general,\
created_from_dataset,\
reference,\
reserve1,\
reserve2,\
reserve3,\
meaning_attribute1,\
meaning_attribute2,\
meaning_attribute3,\
meaning_attribute4,\
meaning_attribute5,\
meaning_attribute6,\
meaning_attribute7,\
meaning_attribute8,\
meaning_attribute9,\
meaning_attribute10,\
meaning_attribute11,\
meaning_attribute12,\
meaning_attribute13,\
meaning_attribute14,\
meaning_attribute15) Values(\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

SQLI = "INSERT INTO classification_items (\
classification_id,\
parent_id,\
description,\
reference,\
attribute1_oto,\
attribute2_oto,\
attribute3_oto,\
attribute4_oto,\
attribute5_anc,\
attribute6_anc,\
attribute7_anc,\
attribute8_anc,\
attribute9_anc,\
attribute10_anc,\
attribute11_anc,\
attribute12_anc,\
attribute13_anc,\
attribute14_anc,\
attribute15_anc) Values(\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

for m in range(0,len(ClassList)):
    ClassFilePath = path.join(path.join(IEDC_Paths.Class_ItemPath,'manual_insert'),ClassList[m] + '.xlsx')
    ClassFile = openpyxl.load_workbook(ClassFilePath, data_only=True)
    ClassDefsheet = ClassFile['Definition']

    C_id       = int(ClassDefsheet.cell(2,2).value)
    C_name     = ClassDefsheet.cell(3,2).value
    C_Dim      = int(ClassDefsheet.cell(4,2).value)
    OtherItems = [] 
    for n in range(0,24):
        if ClassDefsheet.cell(5+n,2).value == 'True':
            OtherItems.append(1)
        elif ClassDefsheet.cell(5+n,2).value == 'False':
            OtherItems.append(0)                        
        elif ClassDefsheet.cell(5+n,2).value != '':
            OtherItems.append(ClassDefsheet.cell(5+n,2).value)            
        else:
            OtherItems.append(None)

    # Define classification
    cur.execute(SQLD,(C_id,C_name,C_Dim,OtherItems[0],OtherItems[1],OtherItems[2],OtherItems[3],OtherItems[4],OtherItems[5],OtherItems[6],OtherItems[7],\
                         OtherItems[8],OtherItems[9],OtherItems[10],OtherItems[11],OtherItems[12],OtherItems[13],OtherItems[14],OtherItems[15],\
                         OtherItems[16],OtherItems[17],OtherItems[18],OtherItems[19],OtherItems[20],OtherItems[21],OtherItems[22],OtherItems[23]))

    # Read items, take attribute1_oto as reference:
    ClassItemsheet = ClassFile['Items']
    n = 1
    while ClassItemsheet.cell(n+1,6).value:
        Items = []
        for o in range(0,18):
            Items.append(ClassItemsheet.cell(n+1,o+3).value)            
        # insert classification item                
        cur.execute(SQLI,(C_id,Items[0],Items[1],Items[2],Items[3],Items[4],Items[5],Items[6],Items[7],Items[8],Items[9],Items[10],\
                         Items[11],Items[12],Items[13],Items[14],Items[15],Items[16],Items[17]))
        n +=1
        
        
        
        
        
# Delete existing classifications:
#cur.execute("DELETE FROM classification_items WHERE classification_id = 38") 
#cur.execute("DELETE FROM classification_definition WHERE id = 38") 
#cur.execute("SELECT * FROM classification_definition WHERE id = 25")
#for row in cur:
#    print(row)

#            
#
## Check
#Value = 27
#cur.execute("SELECT COUNT(*) FROM classification_items WHERE classification_id =%s",(Value))
#for row in cur:
#    print(row)    
#
#cur.execute("SELECT * FROM classification_items WHERE classification_id =%s",(Value))
#for row in cur:
#    print(row)  
#
#Value = 60 
#cur.execute("DELETE FROM classification_items WHERE classification_id =%s",(Value))    
#cur.execute("DELETE FROM classification_definition WHERE id =%s",(Value))       

# Close connection
cur.close()
conn.close()


