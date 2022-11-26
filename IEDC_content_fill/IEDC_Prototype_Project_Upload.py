# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

@author: spauliuk
"""

import pymysql
import datetime
import numpy as np
import openpyxl
from datetime import date, timedelta

import IEDC_PW
import IEDC_Paths

def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1

# Define mySQL command for dataset insertion
SQL = "INSERT INTO projects (\
project_name,\
data_categories,\
data_types,\
data_layers,\
process_scope,\
process_resolution,\
product_scope,\
product_resolution,\
material_scope,\
material_resolution,\
regional_scope,\
regional_resolution,\
temporal_scope,\
temporal_resolution,\
description,\
keywords,\
comment,\
type_of_source,\
project_license,\
main_author,\
project_link,\
project_report,\
suggested_citation,\
submission_date,\
submitting_user,\
reserve1,\
reserve2,\
reserve3) Values(\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,\
%s,%s,%s,%s,%s,%s,%s,%s)"

#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

cur.execute("SELECT table_name, Auto_increment FROM information_schema.tables WHERE table_schema = DATABASE()")    
for row in cur:
    print(row)

# Read ancillary table items:
cur.execute("SELECT id,name FROM licences")
Tuples = cur.fetchall()
DLicen = [x[1] for x in Tuples] 

cur.execute("SELECT id,name FROM source_type")
Tuples = cur.fetchall()
DSource = [x[1] for x in Tuples] 

cur.execute("SELECT id,name FROM users")
Tuples = cur.fetchall()
DUsers = [x[1] for x in Tuples] 

# Read datasets
TOCFile  = openpyxl.load_workbook(IEDC_Paths.DataSetPath + 'IEDC_Prototype_Datasets_Batch1_Upload_MASTER.xlsx', data_only=True)
TOC = TOCFile['Projects']

Offset = 2 # start with Project 3: RECC 
No_Pr  = 1

# loop over datasets
for m in range(Offset,Offset + No_Pr):
    # Define default data:
    D = [] # Data items list
    for n in range(0,29):
        D.append(None) # default
    # loop over items in dataset:

    # 0 ID: auto_increment
    D[1] = TOC.cell(5,m +5).value #1: name    
    D[2] = TOC.cell(8,m +5).value # 2: data categories
    D[3] = TOC.cell(9,m +5).value # 3: data types
    D[4] = TOC.cell(10,m +5).value # 4: data layers
    
    for n in range(5,18):
        D[n]   = TOC.cell(n +4,m +5).value # system location description, dataset description, keywords

    if TOC.cell(22,m +5).value is not None:
        D[18]      = DSource.index(TOC.cell(22,m +5).value) +1  # 18: data source
    if TOC.cell(23,m +5).value is not None:
        D[19]      = DLicen.index(TOC.cell(23,m +5).value) +1   # 19: License

    D[20]      = TOC.cell(24,m +5).value # 20: main author
    D[21]      = TOC.cell(25,m +5).value # 21: link
    D[22]      = TOC.cell(26,m +5).value # 22: report
    D[23]      = TOC.cell(27,m +5).value # 23: citation
        
    D[24]   = TOC.cell(28,m +5).value # submission date
    D[25]   = DUsers.index(TOC.cell(29,m +5).value) +1 # 25: User

    cur.execute(SQL,(D[1],D[2],D[3],D[4],D[5],D[6],D[7],D[8],D[9],D[10],\
                      D[11],D[12],D[13],D[14],D[15],D[16],D[17],D[18],D[19],D[20],\
                      D[21],D[22],D[23],D[24],D[25],D[26],D[27],D[28]))

# Close connection
cur.close()
conn.close()


