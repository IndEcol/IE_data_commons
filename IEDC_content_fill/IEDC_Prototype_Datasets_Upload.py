# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

IEDC_Prototype_Datasets_Upload.py

@author: spauliuk
"""

import pymysql
import datetime
import openpyxl

import IEDC_PW
import IEDC_Paths



# Define mySQL command for dataset insertion
SQL = "INSERT INTO datasets (\
dataset_name,\
dataset_version,\
datagroup_id,\
data_category,\
data_type,\
data_layer,\
process_scope,\
process_resolution,\
product_scope,\
product_resolution,\
material_scope,\
material_resolution,\
regional_scope,\
regional_resolution,\
temporal_scope,\
temporal_resolution,\
description,\
keywords,\
data_provenance,\
dataset_size,\
comment,\
aspect_1,\
aspect_1_classification,\
aspect_2,\
aspect_2_classification,\
aspect_3,\
aspect_3_classification,\
aspect_4,\
aspect_4_classification,\
aspect_5,\
aspect_5_classification,\
aspect_6,\
aspect_6_classification,\
aspect_7,\
aspect_7_classification,\
aspect_8,\
aspect_8_classification,\
aspect_9,\
aspect_9_classification,\
aspect_10,\
aspect_10_classification,\
aspect_11,\
aspect_11_classification,\
aspect_12,\
aspect_12_classification,\
tupel_notation,\
semantic_string_example,\
semantic_string_general,\
type_of_source,\
project_license,\
main_author,\
dataset_link,\
dataset_format,\
project_report,\
suggested_citation,\
visible,\
access_date,\
submission_date,\
submitting_user,\
dataset_conversion_info,\
review_date,\
review_user,\
review_comment,\
reserve1,\
reserve2,\
reserve3,\
reserve4,\
reserve5\
) Values(\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

cur.execute("SELECT table_name, Auto_increment FROM information_schema.tables WHERE table_schema = DATABASE()")    
for row in cur:
    print(row)

## Read ancillary tables:

DClassDefsid = []
DClassDefsD  = []
cur.execute("SELECT id, dimension FROM classification_definition ORDER BY id")
for row in cur:
    DClassDefsid.append(row[0])
    DClassDefsD.append(row[1])

DAspects  = []
DAspectsD = []
cur.execute("SELECT aspect,dimension FROM aspects")
for row in cur:
    DAspects.append(row[0])
    DAspectsD.append(row[1])
    
# Read dataset descriptions
TOCFile  = openpyxl.load_workbook(IEDC_Paths.DataSetPath + 'IEDC_Prototype_Datasets_Batch1_Upload_MASTER.xlsx')
TOC = TOCFile['DataSets_Inventory']

Offset = 114 # entry of row 74 of column of first dataset to be entered.
No_DS = 1

# loop over datasets
for m in range(Offset,Offset + No_DS):
    print(m)    
    # Define default data:
    D = [] # Data items list
    for n in range(0,69):
        D.append(None)
    # loop over items in dataset:

    # 0 ID: auto_increment
    D[1] = TOC.cell(5,m +5).value #1: name
    if  TOC.cell(6,m +5).value != 'NULL': #2: version
        D[2] = TOC.cell(6,m +5).value
    if  TOC.cell(7,m +5).value != 'NULL': #3: datagroup
        D[3] = TOC.cell(7,m +5).value
    
    D[4] = int(TOC.cell(8,m +5).value) # 4: data category
    # 5: data type
    cur.execute("SELECT id FROM types WHERE name = %s ",TOC.cell(9,m +5).value)
    for row in cur:
        D[5] = row[0]
     # 6: data layer
    cur.execute("SELECT id FROM layers WHERE name = %s ",TOC.cell(10,m +5).value)
    for row in cur:
        D[6] = row[0]
        
    for n in range(7,19):
        D[n]   = TOC.cell(n +4,m +5).value # system location description, dataset description, keywords

     # 19: data provenance
    cur.execute("SELECT id FROM provenance WHERE name = %s ",TOC.cell(23,m +5).value)
    for row in cur:
        D[19] = row[0]  

    D[20]      = int(TOC.cell(24,m +5).value) # 20: dataset size
    D[21]      = TOC.cell(25,m +5).value # 21: comment
    
    # Read aspects
    for n in range(0,12):
        a = 2*n   + 22 # aspect index
        c = 2*n+1 + 22 # classification index
        if TOC.cell(a +4,m +5).value != 'none':
            D[a]      = DAspects.index(TOC.cell(a +4,m +5).value) +1
            if TOC.cell(c +4,m +5).value == 'custom':
                D[c]  = 0
            else:
                D[c]  =  int(TOC.cell(c +4,m +5).value)
                if DClassDefsD[DClassDefsid.index(D[c])] !=  DAspectsD[D[a] -1]:
                    print('Aspect and classification mismatch for dataset %s and aspect %s.'  %((m), n))

    D[46]      = TOC.cell(50,m +5).value # 46: tuple notation
    D[47]      = TOC.cell(51,m +5).value # 47: semantic string example
    D[48]      = TOC.cell(52,m +5).value # 48: semantic string general
    # 49: data source
    cur.execute("SELECT id FROM source_type WHERE name = %s ",TOC.cell(53,m +5).value)
    for row in cur:
        D[49] = row[0]     
     # 50: License
    cur.execute("SELECT id FROM licences WHERE name = %s ",TOC.cell(54,m +5).value)
    for row in cur:
        D[50] = row[0]   
        
    for n in range(51,57):
        D[n]   = TOC.cell(n +4,m +5).value  
    
    if TOC.cell(58 +3,m +5).value != 'NULL':
        D[57]   = TOC.cell(58 +3,m +5).value
    D[58]   = TOC.cell(59 +3,m +5).value
     # 59: User
    cur.execute("SELECT id FROM users WHERE name = %s ",TOC.cell(63,m +5).value)
    for row in cur:
        D[59] = row[0] 
    D[60]   = TOC.cell(60 +4,m +5).value 

    cur.execute(SQL,(D[1],D[2],D[3],D[4],D[5],D[6],D[7],D[8],D[9],D[10],\
                    D[11],D[12],D[13],D[14],D[15],D[16],D[17],D[18],D[19],D[20],\
                    D[21],D[22],D[23],D[24],D[25],D[26],D[27],D[28],D[29],D[30],\
                    D[31],D[32],D[33],D[34],D[35],D[36],D[37],D[38],D[39],D[40],\
                    D[41],D[42],D[43],D[44],D[45],D[46],D[47],D[48],D[49],D[50],\
                    D[51],D[52],D[53],D[54],D[55],D[56],D[57],D[58],D[59],D[60],\
                    D[61],D[62],D[63],D[64],D[65],D[66],D[67],D[68]))


    
    
# Close connection
cur.close()
conn.close()



