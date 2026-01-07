# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

@author: spauliuk
"""

import pymysql
import openpyxl

import IEDC_PW

# Define mySQL command for data item insertion:
SQL = "INSERT INTO data (\
dataset_id,\
aspect1,\
aspect2,\
aspect3,\
aspect4,\
aspect5,\
aspect6,\
aspect7,\
aspect8,\
aspect9,\
aspect10,\
aspect11,\
aspect12,\
value,\
unit_nominator,\
unit_denominator,\
stats_array_1,\
stats_array_2,\
stats_array_3,\
stats_array_4,\
comment,\
reserve1,\
reserve2,\
reserve3\
) Values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

# PART 1: Parse data template
DF  = openpyxl.load_workbook('C:\\Users\\Stefan Pauliuk\\FILES\\ARBEIT\\PROJECTS\Database\\IE_DataCommons\\NEW\\NEW_TABLE\\3_MC_Global_3051_Buildings_Database_LIU_2025.xlsx')
DC = DF['Cover']
DD = DF['Data']
dataset_id       = 468
unit_nominator   = 2
unit_denominator = 6

# PART 2: Loop over 3051 rows and check data
for mr in range(0,3051): # 3051 for full dataset
    print(mr)
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(13,DD.cell(mr+2,1).value))
    commodity_ID       = cur.fetchall()[0][0]
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(13,DD.cell(mr+2,2).value))
    product_type_ID    = cur.fetchall()[0][0]
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(2, DD.cell(mr+2,3).value))
    region_ID          = cur.fetchall()[0][0]
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(11,DD.cell(mr+2,4).value))
    country_subunit_ID = cur.fetchall()[0][0]
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(3, DD.cell(mr+2,5).value))
    age_cohort_ID      = cur.fetchall()[0][0]
    cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(6, DD.cell(mr+2,6).value))
    process_ID         = cur.fetchall()[0][0]
    comment            = DD.cell(mr+2,38).value
    # inner loop over the 31 materials
    for mm in range(0,31):
        cur.execute("SELECT id FROM classification_items WHERE classification_id = %s AND attribute1_oto = %s",(4, DD.cell(1,mm+7).value))
        material_ID    = cur.fetchall()[0][0]
        if DD.cell(mr+2,mm+7).value is not None: # insert, discard empty cells
            print(DD.cell(mr+2,mm+7).value)
            MC_Value = DD.cell(mr+2,mm+7).value
            # Set uncertainty string:
            U1 = 1
            U2 = None
            U3 = None
            U4 = None
            # Add data into db:
            # Deactivated after sucessful upload to prevent accidential use
            # cur.execute(SQL,(dataset_id,commodity_ID,product_type_ID,region_ID,country_subunit_ID,age_cohort_ID,process_ID,material_ID,None,None,None,None,None,MC_Value,\
            #               unit_nominator,unit_denominator,U1,U2,U3,U4,comment,None,None,None))

# Close connection
cur.close()
conn.close()

# Delete incomplete uploads:
# cur.execute("DELETE FROM data WHERE dataset_id = 468")

# End



