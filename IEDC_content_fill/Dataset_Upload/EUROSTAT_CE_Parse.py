# -*- coding: utf-8 -*-
"""
Parse and reformat downloaded EuroStat data files for upload to iedc

Created on Sat Aug 26 11:00:17 2023

@author: spauliuk

with input from https://gist.github.com/gka/5443597 

"""

import csv
import openpyxl

def pC_Waste_parse(fn):
    csv_in = csv.reader(open(fn), delimiter='\t')
    head = list(map(lambda k: k.strip(), csv_in.__next__()))
    dim = list(map(lambda k: k.strip(), head[0].split(',')))
    dim[-1], h = dim[-1].split('\\')
    country  = []
    countryc = []
    year     = []
    value    = []
    comment  = []
    for row in csv_in: # parse rows with data
        # explode first column
        rdim = row[0].split(',')
        for i in range(1, len(row[1:])+1):
            country.append(CL_iedc[CL_off.index(rdim[-1])])
            countryc.append(rdim[-1])
            year.append(int(head[i]))
            # check for float value:
            values = row[i].split(' ') # some values contain comments, separated from the numerical value by a blank space
            value.append(get_float(values[0]))
            if values[1]:
                comment.append('Flag set by EUROSTAT: ' + values[1])
            else:
                comment.append(None)
    return country, countryc, year, value, comment

def get_float(s):
    try:
        return float(s)
    except:
        return None

# List of official EUROSTAT to letter country codes, see https://ec.europa.eu/eurostat/statistics-explained/index.php?title=Glossary:Country_codes
CL_off = ['AT',
 'BE',
 'BG',
 'CY',
 'CZ',
 'DE',
 'DK',
 'EE',
 'EL',
 'ES',
 'EU27_2020',
 'FI',
 'FR',
 'HR',
 'HU',
 'IE',
 'IS',
 'IT',
 'LI',
 'LT',
 'LU',
 'LV',
 'MT',
 'NL',
 'NO',
 'PL',
 'PT',
 'RO',
 'SE',
 'SI',
 'SK',
 'UK',
 'AL',
 'BA',
 'CH',
 'ME',
 'MK',
 'RS',
 'TR',
 'XK',
 'EU28']

# Corresponding iedc class. 2 attr. 1 country names list:
CL_iedc = ['Austria',
 'Belgium',
 'Bulgaria',
 'Cyprus',
 'Czech Republic',
 'Germany',
 'Denmark',
 'Estonia',
 'Greece',
 'Spain',
 'EU27',
 'Finland',
 'France',
 'Croatia',
 'Hungary',
 'Ireland',
 'Iceland',
 'Italy',
 'Liechtenstein',
 'Lithuania',
 'Luxembourg',
 'Latvia',
 'Malta',
 'Netherlands',
 'Norway',
 'Poland',
 'Portugal',
 'Romania',
 'Sweden',
 'Slovenia',
 'Slovakia',
 'United Kingdom',
 'Albania',
 'Bosnia and Herzegovina',
 'Switzerland',
 'Montenegro',
 'Macedonia, FYR',
 'Serbia',
 'Turkey',
 'Kosovo',
 'EU28']

waste_list = ['TOTAL',
'PAPER',
'PLAST',
'RUB',
'WD',
'TEXT',
'GLAS',
'ORG',
'MIN',
'MT_FER',
'MT_NFER_PRE',
'MT_NFER_CAN',
'MT_NFER_OTH',
'NSP']

waste_list_clear = ['solid waste (all types)',
'waste paper and cardboard',
'waste plastic',
'waste rubber',
'wood waste',
'waste textiles',
'waste glass',
'waste, organic',
'mineralic waste',
'ferrous metal waste',
'non-ferrous metal waste, precious metals only',
'non-ferrous metal waste, Cu, Al, and Ni only',
'non-ferrous metal waste, other than prec. metals and Cu/Al/Ni',
'waste, not specified']

mat_list = ['CROC',
'AL',
'BE',
'BI',
'CO',
'CU',
'DY',
'GA',
'GE',
'GP',
'IN',
'FE',
'PB',
'LST',
'LI',
'MG',
'MO',
'RUBB',
'ND',
'NI',
'PD',
'PT',
'PR',
'SAPE',
'TA',
'TE',
'TI',
'V',
'Y',
'ZN']
 
mat_list_clear = ['Aggregates - crushed rock, other sands (not silica), pebbles, gravel, bitumen additives',
'aluminium',
'beryllium',
'bismuth',
'cobalt',
'copper',
'dysprosium',
'gallium',
'germanium',
'gypsum',
'indium',
'iron',
'lead',
'limestone',
'lithium',
'magnesium',
'molybdenum',
'natural rubber',
'neodymium',
'nickel',
'palladium',
'platinum',
'praseodymium',
'sapele wood',
'tantalum',
'tellurium',
'titanium',
'vanadium',
'yttrium',
'zinc']

w_list_label = ['W1501','W150101','W150102','W150103','W150104','W150107']
w_list_name  = ['packaging waste',
'paper and cardboard packaging waste',
'plastic packaging waste',
'wooden packaging waste',
'metallic packaging waste',
'glass packaging waste']


########## prepare excel file for per capita waste generation ##########
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'Cover'
ws2 = book.create_sheet('Data')
ws2.cell(row=1, column=1).value = 'region'
ws2.cell(row=1, column=2).value = 'origin_process'
ws2.cell(row=1, column=3).value = 'destination_process'
ws2.cell(row=1, column=4).value = 'material'
ws2.cell(row=1, column=5).value = 'time'
ws2.cell(row=1, column=6).value = 'value'
ws2.cell(row=1, column=7).value = 'unit nominator'
ws2.cell(row=1, column=8).value = 'unit denominator'
ws2.cell(row=1, column=9).value = 'stats_array string'
ws2.cell(row=1, column=10).value = 'comment'
for m in range(1,11):
    ws2.cell(row=1, column=m).font = openpyxl.styles.Font(bold=True)

rowi = 1
########## cei_pc040  (Generation of packaging waste per capita)
fn = 'cei_pc040.tsv'
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,672):
    rowi += 1
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'all (entire economy including households and other end-use sectors)'
    ws2.cell(rowi, column=3).value = 'waste management'
    ws2.cell(rowi, column=4).value = 'packaging waste'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = 'kg'
    ws2.cell(rowi, column=8).value = 'yr'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT codes: cei_pc040, W1501, country code: ' + countryc[m] + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT codes: cei_pc040, W1501, country code: ' + countryc[m]
        
########## cei_pc040  (Generation of plastic packaging waste per capita)
fn = 'cei_pc050.tsv'
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,672):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'all (entire economy including households and other end-use sectors)'
    ws2.cell(rowi, column=3).value = 'waste management'
    ws2.cell(rowi, column=4).value = 'plastic packaging waste'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = 'kg'
    ws2.cell(rowi, column=8).value = 'yr'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc050, W150102, country code: ' + countryc[m] + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc050, W150102, country code: ' + countryc[m]

########## cei_pc031  (Generation of municipal waste per capita)
fn = 'cei_pc031.tsv'
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,858):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'housholds, commerce, and other sources of MSW'
    ws2.cell(rowi, column=3).value = 'waste management'
    ws2.cell(rowi, column=4).value = 'municipal solid waste (MSW)'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = 'kg'
    ws2.cell(rowi, column=8).value = 'yr'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc031, country code: ' + countryc[m] + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc031, country code: ' + countryc[m]

########## cei_pc034  (Total waste generation per capita)
fn = 'cei_pc034.tsv'
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,342):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'all (entire economy including households and other end-use sectors)'
    ws2.cell(rowi, column=3).value = 'waste management'
    ws2.cell(rowi, column=4).value = 'solid waste (all types)'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = 'kg'
    ws2.cell(rowi, column=8).value = 'yr'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc034, country code: ' + countryc[m] + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_pc034, country code: ' + countryc[m]
        
        
########## save excel file ##########
book.save('6_PCF_EUROSTAT_per_capita_waste_flows_parse.xlsx')




########## Waste trade data #########
fn = 'cei_srm020.tsv' # last update: 13-02-2023
csv_in = csv.reader(open(fn), delimiter='\t')
head = list(map(lambda k: k.strip(), csv_in.__next__()))
dim = list(map(lambda k: k.strip(), head[0].split(',')))
dim[-1], h = dim[-1].split('\\')
material = []
flowtype = []
unit     = []
country  = []
countryc = []
year     = []
value    = []
comment  = []
for row in csv_in: # parse rows with data
    # explode first column
    rdim = row[0].split(',')
    for i in range(1, len(row[1:])+1):
        if rdim[2] == 'T': # Only physical flows in t
            country.append(CL_iedc[CL_off.index(rdim[-1])])
            countryc.append(rdim[-1])
            material.append(waste_list_clear[waste_list.index(rdim[0])])
            flowtype.append(rdim[1])
            unit.append(rdim[2])
            year.append(int(head[i]))
            # check for float value:
            values = row[i].split(' ') # some values contain comments, separated from the numerical value by a blank space
            value.append(get_float(values[0]))
            if values[1]:
                comment.append('Flag set by EUROSTAT: ' + values[1])
            else:
                    comment.append(None)

book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'Cover'
ws2 = book.create_sheet('Data')
ws2.cell(row=1, column=1).value = 'material'
ws2.cell(row=1, column=2).value = 'origin_region'
ws2.cell(row=1, column=3).value = 'destination_region'
ws2.cell(row=1, column=4).value = 'origin_process'
ws2.cell(row=1, column=5).value = 'destination_process'
ws2.cell(row=1, column=6).value = 'time'
ws2.cell(row=1, column=7).value = 'value'
ws2.cell(row=1, column=8).value = 'unit nominator'
ws2.cell(row=1, column=9).value = 'unit denominator'
ws2.cell(row=1, column=10).value = 'stats_array string'
ws2.cell(row=1, column=11).value = 'comment'
for m in range(1,12):
    ws2.cell(row=1, column=m).font = openpyxl.styles.Font(bold=True)
    
commentstring = 'EUROSTAT code: cei_srm020, '    

rowi = 1    
for m in range(0,21042):
    rowi += 1   
    commentstring = ''
    ws2.cell(rowi, column=1).value = material[m]
    if flowtype[m] == 'IMP_IEU27_2020':
        ws2.cell(rowi, column=2).value = 'Other EU27 (2020) countries' 
        ws2.cell(rowi, column=3).value = country[m]   
        commentstring += 'imports intra-EU27, country code: '
    if flowtype[m] == 'IMP_XEU27_2020':
        ws2.cell(rowi, column=2).value = 'Non EU27 (2020) countries' 
        ws2.cell(rowi, column=3).value = country[m]   
        commentstring += 'imports extra-EU27, country code: '
    if flowtype[m] == 'EXP_XEU27_2020':
        ws2.cell(rowi, column=2).value = country[m]           
        ws2.cell(rowi, column=3).value = 'Non EU27 (2020) countries' 
        commentstring += 'exports extra-EU27, country code: '
    ws2.cell(rowi, column=4).value = 'market for waste'
    ws2.cell(rowi, column=5).value = 'market for waste'
    ws2.cell(rowi, column=6).value = year[m]
    ws2.cell(rowi, column=7).value = value[m]
    ws2.cell(rowi, column=8).value = 't'
    ws2.cell(rowi, column=9).value = 'yr'
    ws2.cell(rowi, column=10).value = None
    if comment[m]:
        ws2.cell(rowi, column=11).value = commentstring + countryc[m] + comment[m]
    else:
        ws2.cell(rowi, column=11).value = commentstring + countryc[m]    
    
    
book.save('1_F_EUROSTAT_waste_trade_parse.xlsx')      



########## Recycling input rate #########
fn = 'cei_srm010.tsv' # last update: 12-05-2023
csv_in = csv.reader(open(fn), delimiter='\t')
head = list(map(lambda k: k.strip(), csv_in.__next__()))
dim = list(map(lambda k: k.strip(), head[0].split(',')))
dim[-1], h = dim[-1].split('\\')
material = []
year     = []
value    = []
comment  = []
for row in csv_in: # parse rows with data
    # explode first column
    rdim = row[0].split(',')
    for i in range(1, len(row[1:])+1):
        material.append(mat_list_clear[mat_list.index(rdim[0])])
        flowtype.append(rdim[1])
        unit.append(rdim[2])
        year.append(int(head[i]))
        # check for float value:
        values = row[i].split(' ') # some values contain comments, separated from the numerical value by a blank space
        value.append(get_float(values[0]))
        if values[1]:
            comment.append('Flag set by EUROSTAT: ' + values[1])
        else:
                comment.append(None)

book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'Cover'
ws2 = book.create_sheet('Data')
ws2.cell(row=1, column=1).value = 'material'
ws2.cell(row=1, column=2).value = 'region'
ws2.cell(row=1, column=3).value = 'process'
ws2.cell(row=1, column=4).value = 'time'
ws2.cell(row=1, column=5).value = 'value'
ws2.cell(row=1, column=6).value = 'unit nominator'
ws2.cell(row=1, column=7).value = 'unit denominator'
ws2.cell(row=1, column=8).value = 'stats_array string'
ws2.cell(row=1, column=9).value = 'comment'
for m in range(1,10):
    ws2.cell(row=1, column=m).font = openpyxl.styles.Font(bold=True)
    

rowi = 1    
for m in range(0,120):
    rowi += 1   
    commentstring = 'EUROSTAT code: cei_srm010'    
    ws2.cell(rowi, column=1).value = material[m]
    ws2.cell(rowi, column=2).value = 'European Union' 
    ws2.cell(rowi, column=3).value = 'manufacturing'
    ws2.cell(rowi, column=4).value = year[m]
    ws2.cell(rowi, column=5).value = value[m]
    ws2.cell(rowi, column=6).value = '%'
    ws2.cell(rowi, column=7).value = '1'
    ws2.cell(rowi, column=8).value = None
    if comment[m]:
        ws2.cell(rowi, column=9).value = commentstring + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=9).value = commentstring
    
    
book.save('6_MIP_EUROSTAT_recycling_input_rate_parse.xlsx')         


########## Circular material use rate #########
fn = 'cei_srm030.tsv' # last update: 24-01-2023
csv_in = csv.reader(open(fn), delimiter='\t')
head = list(map(lambda k: k.strip(), csv_in.__next__()))
dim = list(map(lambda k: k.strip(), head[0].split(',')))
dim[-1], h = dim[-1].split('\\')
country  = []
countryc = []
year     = []
value    = []
comment  = []
for row in csv_in: # parse rows with data
    # explode first column
    rdim = row[0].split(',')
    for i in range(1, len(row[1:])+1):
        country.append(CL_iedc[CL_off.index(rdim[-1])])
        countryc.append(rdim[-1])
        unit.append(rdim[0])
        year.append(int(head[i]))
        # check for float value:
        values = row[i].split(' ') # some values contain comments, separated from the numerical value by a blank space
        value.append(get_float(values[0]))
        if values[1]:
            comment.append('Flag set by EUROSTAT: ' + values[1])
        else:
                comment.append(None)

book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'Cover'
ws2 = book.create_sheet('Data')
ws2.cell(row=1, column=1).value = 'material'
ws2.cell(row=1, column=2).value = 'region'
ws2.cell(row=1, column=3).value = 'process'
ws2.cell(row=1, column=4).value = 'time'
ws2.cell(row=1, column=5).value = 'value'
ws2.cell(row=1, column=6).value = 'unit nominator'
ws2.cell(row=1, column=7).value = 'unit denominator'
ws2.cell(row=1, column=8).value = 'stats_array string'
ws2.cell(row=1, column=9).value = 'comment'
for m in range(1,10):
    ws2.cell(row=1, column=m).font = openpyxl.styles.Font(bold=True)
    

rowi = 1    
for m in range(0,540):
    rowi += 1   
    commentstring = 'EUROSTAT code: cei_srm030'    
    ws2.cell(rowi, column=1).value = 'all materials'
    ws2.cell(rowi, column=2).value = country[m]  
    ws2.cell(rowi, column=3).value = 'all (entire economy)'
    ws2.cell(rowi, column=4).value = year[m]
    ws2.cell(rowi, column=5).value = value[m]
    ws2.cell(rowi, column=6).value = '%'
    ws2.cell(rowi, column=7).value = '1'
    ws2.cell(rowi, column=8).value = None
    if comment[m]:
        ws2.cell(rowi, column=9).value = commentstring + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=9).value = commentstring
        
book.save('6_MIP_EUROSTAT_circular_material_use_rate_parse.xlsx')   



########## prepare excel file for per recycling rates ##########
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'Cover'
ws2 = book.create_sheet('Data')
ws2.cell(row=1, column=1).value = 'region'
ws2.cell(row=1, column=2).value = 'process'
ws2.cell(row=1, column=3).value = 'input_material'
ws2.cell(row=1, column=4).value = 'output_material'
ws2.cell(row=1, column=5).value = 'time'
ws2.cell(row=1, column=6).value = 'value'
ws2.cell(row=1, column=7).value = 'unit nominator'
ws2.cell(row=1, column=8).value = 'unit denominator'
ws2.cell(row=1, column=9).value = 'stats_array string'
ws2.cell(row=1, column=10).value = 'comment'
for m in range(1,11):
    ws2.cell(row=1, column=m).font = openpyxl.styles.Font(bold=True)

rowi = 1
########## cei_wm010  (all waste excluding major mineral waste)
fn = 'cei_wm010.tsv' # last update: 19-01-2023
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,174):
    rowi += 1
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'waste management'
    ws2.cell(rowi, column=3).value = 'all waste excluding major mineral waste'
    ws2.cell(rowi, column=4).value = 'waste for recycling'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = '%'
    ws2.cell(rowi, column=8).value = '1'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm010, country code: ' + countryc[m] + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm010, country code: ' + countryc[m]
        
########## cei_wm011 (muncipal waste)
fn = 'cei_wm011.tsv' # last update: 21-08-2023
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,858):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'waste management'
    ws2.cell(rowi, column=3).value = 'municipal solid waste (MSW)'
    ws2.cell(rowi, column=4).value = 'waste for recycling'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = '%'
    ws2.cell(rowi, column=8).value = '1'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm011, country code: ' + countryc[m] + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm011, country code: ' + countryc[m]

########## cei_wm020 (packaging waste by type of packaging)
fn = 'cei_wm020.tsv' # last update: 21-03-2023
csv_in = csv.reader(open(fn), delimiter='\t')
head = list(map(lambda k: k.strip(), csv_in.__next__()))
dim = list(map(lambda k: k.strip(), head[0].split(',')))
dim[-1], h = dim[-1].split('\\')
country  = []
countryc = []
w_type   = []
w_label  = []
year     = []
value    = []
comment  = []
for row in csv_in: # parse rows with data
    # explode first column
    rdim = row[0].split(',')
    for i in range(1, len(row[1:])+1):
        country.append(CL_iedc[CL_off.index(rdim[-1])])
        countryc.append(rdim[-1])
        w_type.append(w_list_name[w_list_label.index(rdim[0])])
        w_label.append(rdim[0])
        year.append(int(head[i]))
        # check for float value:
        values = row[i].split(' ') # some values contain comments, separated from the numerical value by a blank space
        value.append(get_float(values[0]))
        if values[1]:
            comment.append('Flag set by EUROSTAT: ' + values[1])
        else:
            comment.append(None)

for m in range(0,4032):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'waste management'
    ws2.cell(rowi, column=3).value = w_type[m]
    ws2.cell(rowi, column=4).value = 'waste for recycling'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = '%'
    ws2.cell(rowi, column=8).value = '1'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm020, country code: ' + countryc[m] + ', waste code:' + w_label[m] + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm020, country code: ' + countryc[m] + ', waste code:' + w_label[m]

########## cei_wm060 (waste electrical and electronic equipment (WEEE))
fn = 'cei_wm060.tsv' # last update: 03-03-2023
country, countryc, year, value, comment = pC_Waste_parse(fn)

for m in range(0,512):
    rowi += 1    
    ws2.cell(rowi, column=1).value = country[m]
    ws2.cell(rowi, column=2).value = 'waste management'
    ws2.cell(rowi, column=3).value = 'waste electrical and electronic equipment (WEEE)'
    ws2.cell(rowi, column=4).value = 'waste for recycling'
    ws2.cell(rowi, column=5).value = year[m]
    ws2.cell(rowi, column=6).value = value[m]
    ws2.cell(rowi, column=7).value = '%'
    ws2.cell(rowi, column=8).value = '1'
    ws2.cell(rowi, column=9).value = None
    if comment[m]:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm060, country code: ' + countryc[m] + ', ' + comment[m]
    else:
        ws2.cell(rowi, column=10).value = 'EUROSTAT code: cei_wm060, country code: ' + countryc[m]
        
        
########## save excel file ##########
book.save('4_PY_EUROSTAT_EoL_Recycling_Rates_parse.xlsx')



# 
# The end.
#


