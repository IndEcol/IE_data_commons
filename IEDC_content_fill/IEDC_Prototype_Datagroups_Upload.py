# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

@author: spauliuk
"""

import pymysql
import datetime
import numpy as np
import openpyxl
from datetime import date, timedelta

import IEDC_PW
import IEDC_Paths

def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1

# Define mySQL command for dataset insertion
SQL = "INSERT INTO datagroups (\
datagroup_name,\
datagroup_version,\
project_id,\
data_categories,\
data_types,\
data_layers,\
process_scope,\
process_resolution,\
product_scope,\
product_resolution,\
material_scope,\
material_resolution,\
regional_scope,\
regional_resolution,\
temporal_scope,\
temporal_resolution,\
description,\
keywords,\
system_definition_picture,\
comment,\
type_of_source,\
project_license,\
main_author,\
project_link,\
project_report,\
suggested_citation,\
submission_date,\
submitting_user,\
reserve1,\
reserve2,\
reserve3) Values(\
%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,\
%s,%s,%s,%s,%s,%s,%s,%s)"

#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

cur.execute("SELECT table_name, Auto_increment FROM information_schema.tables WHERE table_schema = DATABASE()")    
for row in cur:
    print(row)

# Read ancillary table items:
cur.execute("SELECT id,project_name FROM projects")
Tuples = cur.fetchall()
DProj  = [x[1] for x in Tuples]     
    
cur.execute("SELECT id,name FROM licences")
Tuples = cur.fetchall()
DLicen = [x[1] for x in Tuples] 

cur.execute("SELECT id,name FROM source_type")
Tuples = cur.fetchall()
DSource = [x[1] for x in Tuples] 

cur.execute("SELECT id,name FROM users")
Tuples = cur.fetchall()
DUsers = [x[1] for x in Tuples] 

# Read datasets
TOCFile  = openpyxl.load_workbook(IEDC_Paths.DataSetPath + 'IEDC_Prototype_Datasets_Batch1_Upload_MASTER.xlsx')
TOC = TOCFile['DataGroups']

Offset = 24 # last datagroup id that was inserted, currently no. 24 (SOUVERAEN database) in March 2025.
No_DG  = 1  # number of new data groups to insert

# loop over datasets
for m in range(Offset,Offset +No_DG):
    # Define default data:
    D = [] # Data items list
    for n in range(0,32):
        D.append(None) # default
    # loop over items in dataset:

    # 0 ID: auto_increment
    D[1] = TOC.cell(5,m +5).value #1: name
    if  TOC.cell(6,m +5).value != 'none': #2: version
        D[2] = str(TOC.cell(6,m +5).value)
    if  TOC.cell(7,m +5).value != 'none': #3: project id        
        D[3] = DProj.index(TOC.cell(7,m +5).value) +1
    
    D[4] = TOC.cell(8,m +5).value # 4: data categories
    D[5] = TOC.cell(9,m +5).value # 5: data types
    D[6] = TOC.cell(10,m +5).value # 6: data layers
    
    for n in range(7,21):
        D[n]   = TOC.cell(n +4,m +5).value # system location description, dataset description, keywords

    D[21]      = DSource.index(TOC.cell(25,m +5).value) +1  # 21: data source
    D[22]      = DLicen.index(TOC.cell(26,m +5).value) +1 # 22: License

    D[23]      = TOC.cell(27,m +5).value # 23: main author
    D[24]      = TOC.cell(28,m +5).value # 24: link
    D[25]      = TOC.cell(29,m +5).value # 25: report
    D[26]      = TOC.cell(30,m +5).value # 26: citation
        
    D[27]   = TOC.cell(31,m +5).value # submission date
    D[28]   = DUsers.index(TOC.cell(32,m +5).value) +1 # 28: User

    cur.execute(SQL,(D[1],D[2],D[3],D[4],D[5],D[6],D[7],D[8],D[9],D[10],\
                      D[11],D[12],D[13],D[14],D[15],D[16],D[17],D[18],D[19],D[20],\
                      D[21],D[22],D[23],D[24],D[25],D[26],D[27],D[28],D[29],D[30],\
                      D[31]))

# Close connection
cur.close()
conn.close()


