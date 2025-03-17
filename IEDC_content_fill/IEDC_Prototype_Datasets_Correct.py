# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 16:57:29 2017

@author: spauliuk
"""

import pymysql
import openpyxl
import datetime
import numpy as np
import xlrd
from datetime import date, timedelta

import IEDC_PW
import IEDC_Paths

def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1


#conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc_review', autocommit=True, charset='utf8')
conn = pymysql.connect(host='www.industrialecology.uni-freiburg.de', port=3306, user=IEDC_PW.IEDC_write_access_user, passwd=IEDC_PW.IEDC_write_access_user_PW, db='iedc', autocommit=True, charset='utf8')

cur = conn.cursor()

## Change datasets 1_UPI to 4_UPI:
#cur.execute("UPDATE datasets SET dataset_name = '4_UPI_USLCI_Aluminum_cold_rolling_at_plant', data_category = 4 WHERE id = 132")
#cur.execute("UPDATE datasets SET dataset_name = '4_UPI_USLCI_Chainsawing_delimbing', data_category = 4 WHERE id = 133")
#cur.execute("UPDATE datasets SET dataset_name = '4_UPI_USLCI_Electricity_lignite_coal_at_power_plant', data_category = 4 WHERE id = 134")
#cur.execute("UPDATE datasets SET dataset_name = '4_UPI_USLCI_Limestone_at_mine', data_category = 4 WHERE id = 135")    
#cur.execute("UPDATE datasets SET dataset_name = '4_UPI_USLCI_steel_liquid_at_plant', data_category = 4 WHERE id = 136")    

## Add datagroup it to dataset 130 (Steel Sankey Cullen 2012)
#cur.execute("UPDATE datasets SET datagroup_id = 8 WHERE id = 130")    

## Add description of pre 1900 steel cycle datasets 202 and 203:
#cur.execute("UPDATE datasets SET comment = 'only nonzero values are reported, no trade data for the years before 1900 were included' WHERE id = 202")    
#cur.execute("UPDATE datasets SET comment = 'only nonzero values are reported, no trade data for the years before 1900 were included' WHERE id = 203")    

## Add description of 4 sector steel consumption dataset 211:
#cur.execute("UPDATE datasets SET comment = 'only nonzero values are reported!' WHERE id = 211")    

## 19.12.2019
## Add system definition to YSTAFDB
#cur.execute("UPDATE datagroups SET system_definition_picture = 'YSTAFDB_WholeSystem.png' WHERE id = 10")    

## 6.1.2020
## Change DOI for all YSTAFDB-datasets to the paper:
#cur.execute("UPDATE datagroups SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 10")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 233")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 234")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 235")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 236")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 237")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 238")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 239")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 240")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 241")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 242")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 243")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 244")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 247")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 248")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 249")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 250")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 251")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 252")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 253")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 254")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 255")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 256")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 257")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 281")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 284")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 285")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 287")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 290")    
#cur.execute("UPDATE datasets SET suggested_citation = 'DOI 10.1038/s41597-019-0085-7' WHERE id = 292")    

#1.8.21
# Delete dataset values for 195 and 196, as they were uploaded with mistakes using old version of iedc_tools. 
# New upload will be done with up to data iedc_tools
#cur.execute("DELETE FROM data WHERE dataset_id = 195") 
#cur.execute("DELETE FROM data WHERE dataset_id = 196")

#3.8.21
'''
Fix mismatch of region class. 2 for attribute4_oto: code 780 is for Trinidad and not Kosovo!
IEDC bug fix: 780 Kosovo and Trinidad before class. 2 fix (now: 10018 for Kosovo)
Bug fix: 780 country code was twice. Check all datasets with that entry and move, if necessary and where applicable: for all aspects: change classfication id 5987 to 6096 (Kosovo->Trinidad)
All global steel cycle (Pauliuk 2013) data are affected.
'''
# cur.execute("SELECT DISTINCT dataset_id FROM data WHERE aspect7 = 5987") # search for possibly wrong entry for Kosovo 3 digit code
# for row in cur:
#     print(row)     
# Dataset ids found:    
#aspect 1: 2: global steel cycle (Pauliuk 2013) data, only has Trinidad and no Kosovo data: change classfication id 5987 to 6096 (Kosovo->Trinidad)
#cur.execute("UPDATE data SET aspect1 = 6096 WHERE aspect1 = 5987 AND dataset_id = 2") 
#aspect 2: 270 ( 6_MIP_SI.POV.GINI_WorldBank_2019 ): OK! Matching was done via attribute3_oto
#aspect 3: None 
#aspect 4: [70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,96,98,100,102,104,106,108,110,211]: global steel cycle (Pauliuk 2013) data, only has Trinidad and no Kosovo data: change classfication id 5987 to 6096 (Kosovo->Trinidad)
# for m in [70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,96,98,100,102,104,106,108,110,211]: # all are global steel (Pauliuk 2013) datasets, has T&T but no Kosovo, must be wrong and is therefore changed:
#     cur.execute("UPDATE data SET aspect4 = 6096 WHERE aspect4 = 5987 AND dataset_id = %s",(m))  
#     print("UPDATE data SET aspect4 = 6096 WHERE aspect4 = 5987 AND dataset_id = %s",(m))  
#aspect 5: [59,66,67,68,69,]: 2_IUS_steel_200R_4Categories // 2_IUS_steel_200R // 2_S_steel_200R_Slag // 2_S_steel_200R_Obsolete // 2_S_steel_200R_Landfills
# for m in [59,66,67,68,69,]: # all are global steel (Pauliuk 2013) datasets, has T&T but no Kosovo, must be wrong and is therefore changed:
#     cur.execute("UPDATE data SET aspect5 = 6096 WHERE aspect5 = 5987 AND dataset_id = %s",(m))  
#     print("UPDATE data SET aspect5 = 6096 WHERE aspect5 = 5987 AND dataset_id = %s",(m))  
#aspect 6: [70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,97,99,101,103,105,107,109,211]: global steel cycle (Pauliuk 2013) data, only has Trinidad and no Kosovo data: change classfication id 5987 to 6096 (Kosovo->Trinidad)
# for m in [70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,97,99,101,103,105,107,109,211]: # all are global steel (Pauliuk 2013) datasets, has T&T but no Kosovo, must be wrong and is therefore changed:
#     cur.execute("UPDATE data SET aspect6 = 6096 WHERE aspect6 = 5987 AND dataset_id = %s",(m))  
#     print("UPDATE data SET aspect6 = 6096 WHERE aspect6 = 5987 AND dataset_id = %s",(m))  
#aspect 7+: None

#counter-search for Trinidad:
# cur.execute("SELECT DISTINCT dataset_id FROM data WHERE aspect7 = 6096") # search for possibly wrong entry for Trinidad 3 digit code
# for row in cur:
#     print(row)     
# Dataset ids found BEFORE change:
#aspect 1: 195,196: correct! Was uploaded with distinct ids for Kosovo and Trinidad, those datasets triggered the whole process. 
#aspect 2: 191: 1_F_NewPassengerVehicleRegistration_AllCountries_OICA_2017: OK! contains data for T&T but not for Kosovo, 194: 2_IUS_Vehicles_AllCountries_OICA_2017: OK! contains data for T&T but not for Kosovo, 270: 6_MIP_SI.POV.GINI_WorldBank_2019: OK! Matching was done via attribute3_oto
#aspect 3+: None

# Dataset ids found BEFORE change:
#aspect 1: 2, 195,196: OK!
#aspect 2: 191: 1_F_NewPassengerVehicleRegistration_AllCountries_OICA_2017: OK! contains data for T&T but not for Kosovo, 194: 2_IUS_Vehicles_AllCountries_OICA_2017: OK! contains data for T&T but not for Kosovo, 270: 6_MIP_SI.POV.GINI_WorldBank_2019: OK! Matching was done via attribute3_oto
#aspect 3: None
#aspect 4: Same as prev. aspect 4 search for Kosovo id
#aspect 5: Same as prev. aspect 5 search for Kosovo id
#aspect 6: Same as prev. aspect 6 search for Kosovo id
#aspect 7+: None

### 25.11.2022 update
#cur.execute("UPDATE datasets SET aspect_2_classification = 77 WHERE id = 302") 
#cur.execute("UPDATE datasets SET aspect_3_classification = 77 WHERE id = 304") 
#cur.execute("UPDATE datasets SET dataset_version = 'as published' WHERE id = 305") 
#cur.execute("UPDATE datasets SET dataset_version = 'v1.0' WHERE id = 300") 
#cur.execute("UPDATE datasets SET datagroup_id = 12 WHERE id = 300") 

# cur.execute("UPDATE datagroups SET system_definition_picture = 'RECC_SysDef_Model_v2_4.png' WHERE id = 12") 
# cur.execute("UPDATE datagroups SET system_definition_picture = 'RECC_SysDef_Model_v2_5.png' WHERE id = 13") 
# cur.execute("UPDATE datagroups SET system_definition_picture = '' WHERE id = 14") 

### Nov. 2023 update
#cur.execute("UPDATE datasets SET dataset_version = 'none' WHERE id = 309") 
#cur.execute("UPDATE datasets SET dataset_version = 'none' WHERE id = 312") 
#cur.execute("DELETE FROM datasets WHERE id = 315") 
#cur.execute("UPDATE datasets SET dataset_version = 'none' WHERE id = 319") 
#cur.execute("UPDATE datasets SET dataset_version = 'none' WHERE id = 320") 

### Nov. 2024 update
#cur.execute("UPDATE datagroups SET project_link = 'https://doi.org/10.1111/jiec.13557' WHERE id = 13") 
#cur.execute("UPDATE datagroups SET project_report = 'https://doi.org/10.1111/jiec.13557' WHERE id = 13")
#cur.execute("UPDATE datagroups SET suggested_citation = 'https://doi.org/10.1111/jiec.13557' WHERE id = 13") 

#Steel stocks and flows 2013 paper: Add comment that all data refer to method (a) as described in the paper.
#cur.execute("UPDATE datasets SET comment = 'all data refer to method (a) as described in the paper' WHERE datagroup_id = 1 AND comment = 'none'") 
#cur.execute("UPDATE datasets SET aspect_4_classification = 14 WHERE id = 362") 

# March 2025:
# Search for mis-spelled classification items    
# for m in range(1,13): 
#     cur.execute("SELECT DISTINCT dataset_id FROM data WHERE aspect" + str(m) + " = 1432") # search for possibly wrong entry for a given classification item
#     for row in cur:
#         print(row) 

#Search for mis-spelled classification items    
# for m in range(1,13): 
#     cur.execute("SELECT DISTINCT id,dataset_id FROM data WHERE aspect" + str(m) + " = 1354") # search for possibly wrong entry for a given classification item
#     for row in cur:
#         print(row) 
#         print(m)
# returns (3160714, 378)
#cur.execute("UPDATE data SET aspect1 = 1455 WHERE aspect1 = 1354 AND dataset_id = 378") 

# update datagroup ids:
#cur.execute("UPDATE datasets SET datagroup_id = 4 WHERE id = 209")     

#Correct projects and datasets, as upload scripts contained a mistake until now:
# Fix iedc.projects type of source, fix iedc.datagroups type of source and license
# Read dataset descriptions
#TOCFile  = openpyxl.load_workbook(IEDC_Paths.DataSetPath + 'IEDC_Prototype_Datasets_Batch1_Upload_MASTER.xlsx')
# TOC = TOCFile['Projects']
# for m in range (5,12):
#     print(TOC.cell(22,m).value)
#     # fetch index for provenance:
#     cur.execute("SELECT id FROM source_type WHERE name = %s ",TOC.cell(22,m).value)
#     for row in cur:
#         provs = row[0]      
#         print(provs)
#         print(m-4)
#     # update entry in projects table:
#     cur.execute("UPDATE projects SET type_of_source = %s WHERE id = %s",(provs,m-4))      
    
#TOC = TOCFile['DataGroups']
# for m in range (5,29):
#     print(TOC.cell(25,m).value)
#     # fetch index for provenance:
#     cur.execute("SELECT id FROM source_type WHERE name = %s ",TOC.cell(25,m).value)
#     for row in cur:
#         provs = row[0]      
#         print(provs)
#         print(m-4)
#     # update entry in datagroups table:
#     cur.execute("UPDATE datagroups SET type_of_source = %s WHERE id = %s",(provs,m-4))    

# for m in range (5,29):
#     print(TOC.cell(26,m).value)
#     # fetch index for licences:
#     cur.execute("SELECT id FROM licences WHERE name = %s ",TOC.cell(26,m).value)
#     for row in cur:
#         lics = row[0]      
#         print(lics)
#         print(m-4)
#     # update entry in datagroups table:
#     cur.execute("UPDATE datagroups SET project_license = %s WHERE id = %s",(lics,m-4)) 
    
# change version from number to string for id 365:
#cur.execute("UPDATE datasets SET dataset_version = 'v2023' WHERE id = 365")  

# 4) close connection
# cur.close()
# conn.close()




#